# Utility functions for expense calculation
import os
from openpyxl import Workbook, load_workbook

def calculate_settlement(participants, expenses):
    balances = {p: 0 for p in participants}
    for exp in expenses:
        involved = [p for p in participants if p not in exp.omitted]
        share = exp.amount / len(involved)
        for p in involved:
            balances[p] -= share
        balances[exp.paid_by] += exp.amount
    return balances


def save_settlement_to_excel(trip_name, participants, expenses, balances, archive_dir):
    # Save using openpyxl directly to avoid heavyweight pandas dependency.
    if not os.path.exists(archive_dir):
        os.makedirs(archive_dir)

    file_path = os.path.join(archive_dir, f'{trip_name}.xlsx')

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    ws_summary.append(['Name', 'Type', 'Amount'])
    for p, bal in balances.items():
        if abs(bal) < 0.01:
            continue
        if bal > 0:
            ws_summary.append([p, 'Receive', round(bal, 2)])
        else:
            ws_summary.append([p, 'Pay', round(-bal, 2)])

    ws_expenses = wb.create_sheet('Expenses')
    ws_expenses.append(['Item', 'Amount', 'Paid By', 'Omitted', 'Notes'])
    for exp in expenses:
        ws_expenses.append([
            exp.item,
            exp.amount,
            exp.paid_by,
            ', '.join(exp.omitted) if exp.omitted else '',
            getattr(exp, 'notes', ''),
        ])

    wb.save(file_path)
    return file_path


def load_trip_from_excel(file_path):
    wb = load_workbook(file_path, data_only=True)

    participants = []
    if 'Summary' in wb.sheetnames:
        ws_summary = wb['Summary']
        # Skip header row.
        for row in ws_summary.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            name = row[0]
            if name and name not in participants:
                participants.append(str(name))

    expenses = []
    if 'Expenses' in wb.sheetnames:
        ws_expenses = wb['Expenses']
        for row in ws_expenses.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            item = row[0] if len(row) > 0 else ''
            amount = row[1] if len(row) > 1 else 0
            paid_by = row[2] if len(row) > 2 else ''
            omitted = row[3] if len(row) > 3 else ''
            notes = row[4] if len(row) > 4 else ''
            expenses.append({
                'Item': str(item) if item is not None else '',
                'Amount': float(amount) if amount is not None else 0.0,
                'Paid By': str(paid_by) if paid_by is not None else '',
                'Omitted': str(omitted) if omitted else '',
                'Notes': str(notes) if notes else '',
            })

    return participants, expenses
